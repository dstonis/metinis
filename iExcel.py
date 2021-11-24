from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, colors
from openpyxl.cell import Cell

import numpy as np
import cv2

from colormap import rgb2hex

failas = load_workbook("alio.xlsx")
sheet = failas.active

vardas = "spalvos.PNG"
img = cv2.imread(vardas, 1)
x, y = len(img), len(img[0])

res = x*y

for X in range(x):
    for Y in range(y):
        a, b, c = img[X, Y]
        kodas = rgb2hex(c, b,
                        a)[1:]
        sheet.cell(X+1, Y+1).fill = PatternFill("solid", fgColor=str(kodas))
        print((X+1),(Y+1), kodas, a, b, c)

        
failas.save("alio.xlsx")
